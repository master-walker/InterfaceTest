#!/usr/bin/env python
#coding=utf-8

'''

common function
'''

from openpyxl import load_workbook

# 获取工作表
def get_sheet(file_path, sheet_title=''):
    workbook = load_workbook(file_path)
    work_sheet = workbook[sheet_title]
    return work_sheet

def get_cell_value():
    ws = get_sheet()
    columns = ws.columns
    for col in columns:
        num = ws.cell(col).value





